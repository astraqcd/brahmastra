import openpyxl
import json

wb = openpyxl.load_workbook('bRAHMASTRA.xlsx')

all_tools = []
all_categories = set()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"PROCESSING SHEET: {sheet_name}")
    print(f"{'='*60}\n")
    
    # Try to find header row
    headers = None
    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
        if row and any(str(cell).lower() in ['tool', 'name', 'url', 'link', 'website'] for cell in row if cell):
            headers = row
            print(f"Found headers at row {row_idx}: {headers}")
            break
    
    # Extract tools
    if headers:
        tool_col = None
        url_col = None
        category_col = None
        desc_col = None
        
        for idx, header in enumerate(headers):
            if header:
                h = str(header).lower().strip()
                if 'tool' in h or 'name' in h:
                    tool_col = idx
                elif 'url' in h or 'link' in h or 'website' in h:
                    url_col = idx
                elif 'categ' in h or 'type' in h:
                    category_col = idx
                elif 'desc' in h or 'about' in h:
                    desc_col = idx
        
        print(f"Tool column: {tool_col}, URL column: {url_col}, Category column: {category_col}")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=row_idx+2, values_only=True), row_idx+2):
            if row and row[tool_col if tool_col is not None else 0]:
                tool_name = row[tool_col] if tool_col is not None else row[0]
                tool_url = row[url_col] if url_col is not None and url_col < len(row) else None
                tool_category = row[category_col] if category_col is not None and category_col < len(row) else sheet_name
                tool_desc = row[desc_col] if desc_col is not None and desc_col < len(row) else ""
                
                if tool_name and str(tool_name).strip():
                    tool_entry = {
                        'name': str(tool_name).strip(),
                        'url': str(tool_url).strip() if tool_url else '',
                        'category': str(tool_category).strip() if tool_category else sheet_name,
                        'description': str(tool_desc).strip() if tool_desc else '',
                        'sheet': sheet_name
                    }
                    all_tools.append(tool_entry)
                    if tool_category:
                        all_categories.add(str(tool_category).strip())
                    
                    print(f"  • {tool_entry['name']} ({tool_entry['category']})")
    else:
        # No headers found, try to extract as simple list
        print("No clear headers found, extracting as simple list...")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and row[0] and str(row[0]).strip():
                tool_name = str(row[0]).strip()
                tool_url = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                
                # Skip obvious headers
                if tool_name.lower() not in ['tool', 'name', 'url', 'link', 'website', 'category']:
                    tool_entry = {
                        'name': tool_name,
                        'url': tool_url,
                        'category': sheet_name,
                        'description': '',
                        'sheet': sheet_name
                    }
                    all_tools.append(tool_entry)
                    all_categories.add(sheet_name)
                    print(f"  • {tool_name}")

print(f"\n{'='*60}")
print(f"SUMMARY")
print(f"{'='*60}")
print(f"Total tools found: {len(all_tools)}")
print(f"Total categories: {len(all_categories)}")
print(f"Categories: {sorted(all_categories)}")

# Save to JSON for easy import
output = {
    'tools': all_tools,
    'categories': sorted(all_categories)
}

with open('extracted_tools.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print(f"\n✅ Data saved to extracted_tools.json")
